from flask import Flask, request, jsonify, render_template, send_file
from flask_cors import CORS
import psycopg2
from psycopg2.extras import RealDictCursor
import os
import base64
import google.generativeai as genai
from datetime import datetime
import json
import re
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from ocr_utils import OCRProcessor
from utils import clean_option, validate_image_file, sanitize_filename, validate_roll_number, calculate_percentage, extract_marks_from_text, clean_multiple_options, validate_multiple_options, calculate_partial_marks, parse_options_string

app = Flask(__name__)
CORS(app)

DATABASE_URL = os.environ.get('DATABASE_URL', "postgresql://neondb_owner:npg_Q0NxwyzY5rBt@ep-cold-wildflower-a8n97rw3-pooler.eastus2.azure.neon.tech/neondb?sslmode=require&channel_binding=require")

# Load all API keys for backup system
GEMINI_API_KEY = os.environ.get('GEMINI_API_KEY', "AIzaSyAgKFZq183p04eeHGQThTs7t2eAvhwFzJ4")
GEMINI_API_KEY_BACKUP_1 = os.environ.get('GEMINI_API_KEY_BACKUP_1', "AIzaSyA2rKi4X3LyiRYOnE70ZS6P-BeA8d-6HkM")
GEMINI_API_KEY_BACKUP_2 = os.environ.get('GEMINI_API_KEY_BACKUP_2', "AIzaSyBjtiUdljU6qec1m0X9Sclb4bFYiNkISoY")
GEMINI_API_KEY_BACKUP_3 = os.environ.get('GEMINI_API_KEY_BACKUP_3', "AIzaSyBbK9a8x80b8qV6Odj9x-bZTIZLb7zwkOc")
GEMINI_API_KEY_BACKUP_4 = os.environ.get('GEMINI_API_KEY_BACKUP_4', "AIzaSyDsOFThZxJI5PgO3iFOWX4Kk6W41KUz890")
GEMINI_API_KEY_BACKUP_5 = os.environ.get('GEMINI_API_KEY_BACKUP_5', "AIzaSyCmbOvwgGCJOch2TzpCFvHGbj0tTsdwQVk")

# Initialize API Key Manager (this will configure Gemini automatically)
from api_key_manager import get_api_manager
api_manager = get_api_manager()

# Only print on first load, not on Flask debug restart
if not os.environ.get('WERKZEUG_RUN_MAIN'):
    print(f"🔑 Flask app initialized with {api_manager.get_status()['total_keys']} API keys for automatic rotation")
ocr_processor = OCRProcessor()  # No need to pass API key, uses API manager

def get_db_connection():
    """Get database connection with improved error handling"""
    try:
        conn = psycopg2.connect(
            DATABASE_URL, 
            cursor_factory=RealDictCursor,
            connect_timeout=10,
            application_name='flask-ocr-grading'
        )
        return conn
    except Exception as e:
        print(f"Database connection error: {e}")
        return None

def init_database():
    """Initialize database tables with better error handling"""
    conn = get_db_connection()
    if not conn:
        print("Failed to connect to database for initialization")
        return False
    
    try:
        cursor = conn.cursor()
        
        # Create question_papers table
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS question_papers (
                id SERIAL PRIMARY KEY,
                paper_name VARCHAR(255) NOT NULL,
                total_questions INTEGER NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        # Create correct_answers table
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS correct_answers (
                id SERIAL PRIMARY KEY,
                paper_id INTEGER REFERENCES question_papers(id) ON DELETE CASCADE,
                question_number INTEGER NOT NULL,
                correct_option VARCHAR(10) NOT NULL,
                marks INTEGER DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        # Create student_submissions table
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS student_submissions (
                id SERIAL PRIMARY KEY,
                paper_id INTEGER REFERENCES question_papers(id) ON DELETE CASCADE,
                roll_number VARCHAR(50) NOT NULL,
                student_name VARCHAR(255),
                submitted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        # Create student_answers table
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS student_answers (
                id SERIAL PRIMARY KEY,
                submission_id INTEGER REFERENCES student_submissions(id) ON DELETE CASCADE,
                question_number INTEGER NOT NULL,
                selected_option VARCHAR(10) NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        # Create results table
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS results (
                id SERIAL PRIMARY KEY,
                submission_id INTEGER REFERENCES student_submissions(id) ON DELETE CASCADE,
                total_questions INTEGER NOT NULL,
                correct_answers INTEGER NOT NULL,
                total_marks INTEGER NOT NULL,
                percentage DECIMAL(5,2) NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        # Create indexes for better performance
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_correct_answers_paper_id ON correct_answers(paper_id)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_student_submissions_roll ON student_submissions(roll_number)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_student_answers_submission ON student_answers(submission_id)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_results_submission ON results(submission_id)")
        
        # Add new columns for multiple correct options support (migration)
        try:
            # Check if correct_options column exists
            cursor.execute("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name='correct_answers' AND column_name='correct_options';
            """)
            
            if not cursor.fetchone():
                # Add the correct_options column (array to store multiple options)
                cursor.execute("ALTER TABLE correct_answers ADD COLUMN correct_options TEXT[];")
                print("✅ Added correct_options column to correct_answers table")
            
            # Check if selected_options column exists
            cursor.execute("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name='student_answers' AND column_name='selected_options';
            """)
            
            if not cursor.fetchone():
                # Add the selected_options column (array to store multiple selected options)
                cursor.execute("ALTER TABLE student_answers ADD COLUMN selected_options TEXT[];")
                print("✅ Added selected_options column to student_answers table")
            
            # Check if question_type column exists
            cursor.execute("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name='correct_answers' AND column_name='question_type';
            """)
            
            if not cursor.fetchone():
                # Add question_type column to distinguish single vs multiple correct answers
                cursor.execute("ALTER TABLE correct_answers ADD COLUMN question_type VARCHAR(20) DEFAULT 'single';")
                print("✅ Added question_type column to correct_answers table")
            
            # Create indexes for better performance
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_correct_answers_type ON correct_answers(question_type);")
            
            # Check if images_count column exists in student_submissions
            cursor.execute("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name='student_submissions' AND column_name='images_count';
            """)
            
            if not cursor.fetchone():
                # Add images_count column to track multiple image submissions
                cursor.execute("ALTER TABLE student_submissions ADD COLUMN images_count INTEGER DEFAULT 1;")
                print("✅ Added images_count column to student_submissions table")
            
            # Check if paper_info columns exist in student_submissions
            paper_info_columns = [
                ('extracted_paper_name', 'TEXT'),
                ('extracted_subject', 'TEXT'),
                ('extracted_date', 'TEXT'),
                ('extracted_duration', 'TEXT'),
                ('extracted_total_marks', 'TEXT'),
                ('extracted_class_grade', 'TEXT')
            ]
            
            for column_name, column_type in paper_info_columns:
                cursor.execute(f"""
                    SELECT column_name 
                    FROM information_schema.columns 
                    WHERE table_name='student_submissions' AND column_name='{column_name}';
                """)
                
                if not cursor.fetchone():
                    cursor.execute(f"ALTER TABLE student_submissions ADD COLUMN {column_name} {column_type};")
                    print(f"✅ Added {column_name} column to student_submissions table")
            
            # Migrate existing data - populate correct_options array from correct_option
            cursor.execute("""
                UPDATE correct_answers 
                SET correct_options = ARRAY[correct_option]
                WHERE correct_options IS NULL AND correct_option IS NOT NULL;
            """)
            
            # Migrate existing student answers - populate selected_options array from selected_option
            cursor.execute("""
                UPDATE student_answers 
                SET selected_options = ARRAY[selected_option]
                WHERE selected_options IS NULL AND selected_option IS NOT NULL;
            """)
            
            # Only print on main process, not on Flask debug restart
            if not os.environ.get('WERKZEUG_RUN_MAIN'):
                print("✅ Multiple-correct MCQ migration completed successfully!")
            
        except Exception as migration_error:
            print(f"Migration warning (may be already applied): {migration_error}")
            # Continue with normal initialization even if migration fails
        
        conn.commit()
        cursor.close()
        conn.close()
        # Only print on main process, not on Flask debug restart
        if not os.environ.get('WERKZEUG_RUN_MAIN'):
            print("Database initialized successfully")
        return True
        
    except Exception as e:
        print(f"Database initialization error: {e}")
        if conn:
            conn.rollback()
            conn.close()
        return False

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/status')
def api_status():
    """Get API key manager status"""
    try:
        status = api_manager.get_status()
        return jsonify({
            "success": True,
            "api_status": status,
            "message": f"Using key {status['current_key_index'] + 1} of {status['total_keys']}"
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

@app.route('/api/reset-keys', methods=['POST'])
def reset_api_keys():
    """Reset failed API keys (useful for daily quota reset)"""
    try:
        api_manager.reset_failed_keys()
        status = api_manager.get_status()
        return jsonify({
            "success": True,
            "message": "All API keys reset successfully",
            "api_status": status
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

@app.route('/api/papers', methods=['GET'])
def get_papers():
    """Get all question papers"""
    conn = get_db_connection()
    if not conn:
        return jsonify({"error": "Database connection failed"}), 500
    
    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT qp.*, COUNT(ca.id) as answer_count 
            FROM question_papers qp 
            LEFT JOIN correct_answers ca ON qp.id = ca.paper_id 
            GROUP BY qp.id 
            ORDER BY qp.created_at DESC
        """)
        papers = cursor.fetchall()
        cursor.close()
        conn.close()
        
        return jsonify({"papers": [dict(paper) for paper in papers]})
        
    except Exception as e:
        if conn:
            conn.close()
        return jsonify({"error": f"Failed to fetch papers: {str(e)}"}), 500

@app.route('/api/upload-answer-key', methods=['POST'])
def teacher_upload_ocr():
    """Teacher upload answer key with OCR"""
    try:
        if 'image' not in request.files:
            return jsonify({"error": "No image file provided"}), 400
        
        image_file = request.files['image']
        paper_name = request.form.get('paper_name', '').strip()
        
        if not paper_name:
            return jsonify({"error": "Paper name is required"}), 400
        
        if not validate_image_file(image_file):
            return jsonify({"error": "Invalid image file"}), 400
        
        # Process image with OCR
        image_data = ocr_processor.preprocess_image(image_file)
        if not image_data:
            return jsonify({"error": "Failed to process image"}), 400
        
        # Extract answers using OCR
        ocr_result = ocr_processor.extract_teacher_answers(image_data)
        
        # Validate OCR result
        is_valid, error_msg = ocr_processor.validate_teacher_response(ocr_result)
        if not is_valid:
            return jsonify({"error": f"OCR validation failed: {error_msg}"}), 400
        
        # Save to database
        conn = get_db_connection()
        if not conn:
            return jsonify({"error": "Database connection failed"}), 500
        
        try:
            cursor = conn.cursor()
            
            # Insert question paper
            cursor.execute("""
                INSERT INTO question_papers (paper_name, total_questions) 
                VALUES (%s, %s) RETURNING id
            """, (paper_name, len(ocr_result['answers'])))
            
            paper_id = cursor.fetchone()['id']
            
            # Insert correct answers
            for answer in ocr_result['answers']:
                # Extract marks with multiple fallback strategies
                marks = answer.get('marks', 1)
                
                # If marks is still 1 (default), try to extract from marks_text
                if marks == 1 and 'marks_text' in answer and answer['marks_text']:
                    extracted_marks = extract_marks_from_text(answer['marks_text'])
                    if extracted_marks > 1:  # Only use if we found something meaningful
                        marks = extracted_marks
                
                # Additional fallback: try to extract from any text fields
                if marks == 1:
                    # Check if there's any other text that might contain marks
                    for key, value in answer.items():
                        if isinstance(value, str) and any(char.isdigit() for char in value):
                            potential_marks = extract_marks_from_text(value)
                            if potential_marks > 1:
                                marks = potential_marks
                                break
                
                # Handle both new format (correct_options) and old format (correct_option)
                if 'correct_options' in answer:
                    # New format with multiple options support
                    correct_options = clean_multiple_options(answer['correct_options'])
                    question_type = answer.get('question_type', 'single' if len(correct_options) == 1 else 'multiple')
                    
                    # For backward compatibility, also set correct_option to first option
                    correct_option = correct_options[0] if correct_options else 'A'
                    
                    print(f"DEBUG: Question {answer['question_number']}: marks={marks}, options={correct_options}, type={question_type}")
                    
                    # Try to insert with new schema, fallback to old schema if needed
                    try:
                        cursor.execute("""
                            INSERT INTO correct_answers (paper_id, question_number, correct_option, correct_options, marks, question_type) 
                            VALUES (%s, %s, %s, %s, %s, %s)
                        """, (
                            paper_id, 
                            answer['question_number'], 
                            correct_option,
                            correct_options,
                            marks,
                            question_type
                        ))
                    except Exception as e:
                        # Fallback to old schema if new columns don't exist
                        print(f"Using fallback OCR correct answer insert due to: {e}")
                        cursor.execute("""
                            INSERT INTO correct_answers (paper_id, question_number, correct_option, marks) 
                            VALUES (%s, %s, %s, %s)
                        """, (
                            paper_id, 
                            answer['question_number'], 
                            correct_option,
                            marks
                        ))
                else:
                    # Old format - backward compatibility
                    correct_option = clean_option(answer['correct_option'])
                    
                    print(f"DEBUG: Question {answer['question_number']}: marks={marks}, option={correct_option} (legacy format)")
                    
                    # Try to insert with new schema, fallback to old schema if needed
                    try:
                        cursor.execute("""
                            INSERT INTO correct_answers (paper_id, question_number, correct_option, correct_options, marks, question_type) 
                            VALUES (%s, %s, %s, %s, %s, %s)
                        """, (
                            paper_id, 
                            answer['question_number'], 
                            correct_option,
                            [correct_option],  # Convert to array format
                            marks,
                            'single'
                        ))
                    except Exception as e:
                        # Fallback to old schema if new columns don't exist
                        print(f"Using fallback OCR legacy correct answer insert due to: {e}")
                        cursor.execute("""
                            INSERT INTO correct_answers (paper_id, question_number, correct_option, marks) 
                            VALUES (%s, %s, %s, %s)
                        """, (
                            paper_id, 
                            answer['question_number'], 
                            correct_option,
                            marks
                        ))
            
            conn.commit()
            cursor.close()
            conn.close()
            
            return jsonify({
                "success": True,
                "message": "Answer key uploaded successfully",
                "paper_id": paper_id,
                "total_questions": len(ocr_result['answers']),
                "extracted_answers": f"{len(ocr_result['answers'])} answers extracted"
            })
            
        except Exception as e:
            if conn:
                conn.rollback()
                conn.close()
            return jsonify({"error": f"Database error: {str(e)}"}), 500
            
    except Exception as e:
        return jsonify({"error": f"Upload failed: {str(e)}"}), 500

@app.route('/api/manual-answer-key', methods=['POST'])
def teacher_upload_manual():
    """Teacher upload answer key manually"""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({"error": "No data provided"}), 400
        
        paper_name = data.get('paper_name', '').strip()
        answers = data.get('answers', [])
        
        if not paper_name:
            return jsonify({"error": "Paper name is required"}), 400
        
        if not answers or not isinstance(answers, list):
            return jsonify({"error": "Answers array is required"}), 400
        
        # Save to database
        conn = get_db_connection()
        if not conn:
            return jsonify({"error": "Database connection failed"}), 500
        
        try:
            cursor = conn.cursor()
            
            # Insert question paper
            cursor.execute("""
                INSERT INTO question_papers (paper_name, total_questions) 
                VALUES (%s, %s) RETURNING id
            """, (paper_name, len(answers)))
            
            paper_id = cursor.fetchone()['id']
            
            # Insert correct answers
            for answer in answers:
                # Extract marks with multiple fallback strategies
                marks = answer.get('marks', 1)
                
                # If marks is still 1 (default), try to extract from marks_text
                if marks == 1 and 'marks_text' in answer and answer['marks_text']:
                    extracted_marks = extract_marks_from_text(answer['marks_text'])
                    if extracted_marks > 1:  # Only use if we found something meaningful
                        marks = extracted_marks
                
                # Additional fallback: try to extract from any text fields
                if marks == 1:
                    # Check if there's any other text that might contain marks
                    for key, value in answer.items():
                        if isinstance(value, str) and any(char.isdigit() for char in value):
                            potential_marks = extract_marks_from_text(value)
                            if potential_marks > 1:
                                marks = potential_marks
                                break
                
                # Handle both new format (correct_options) and old format (correct_option)
                if 'correct_options' in answer:
                    # New format with multiple options support
                    correct_options = clean_multiple_options(answer['correct_options'])
                    question_type = answer.get('question_type', 'single' if len(correct_options) == 1 else 'multiple')
                    
                    # For backward compatibility, also set correct_option to first option
                    correct_option = correct_options[0] if correct_options else 'A'
                    
                    print(f"DEBUG: Manual Question {answer['question_number']}: marks={marks}, options={correct_options}, type={question_type}")
                    
                    # Try to insert with new schema, fallback to old schema if needed
                    try:
                        cursor.execute("""
                            INSERT INTO correct_answers (paper_id, question_number, correct_option, correct_options, marks, question_type) 
                            VALUES (%s, %s, %s, %s, %s, %s)
                        """, (
                            paper_id, 
                            answer['question_number'], 
                            correct_option,
                            correct_options,
                            marks,
                            question_type
                        ))
                    except Exception as e:
                        # Fallback to old schema if new columns don't exist
                        print(f"Using fallback manual correct answer insert due to: {e}")
                        cursor.execute("""
                            INSERT INTO correct_answers (paper_id, question_number, correct_option, marks) 
                            VALUES (%s, %s, %s, %s)
                        """, (
                            paper_id, 
                            answer['question_number'], 
                            correct_option,
                            marks
                        ))
                else:
                    # Old format - backward compatibility
                    correct_option = clean_option(answer['correct_option'])
                    
                    print(f"DEBUG: Manual Question {answer['question_number']}: marks={marks}, option={correct_option} (legacy format)")
                    
                    # Try to insert with new schema, fallback to old schema if needed
                    try:
                        cursor.execute("""
                            INSERT INTO correct_answers (paper_id, question_number, correct_option, correct_options, marks, question_type) 
                            VALUES (%s, %s, %s, %s, %s, %s)
                        """, (
                            paper_id, 
                            answer['question_number'], 
                            correct_option,
                            [correct_option],  # Convert to array format
                            marks,
                            'single'
                        ))
                    except Exception as e:
                        # Fallback to old schema if new columns don't exist
                        print(f"Using fallback manual legacy correct answer insert due to: {e}")
                        cursor.execute("""
                            INSERT INTO correct_answers (paper_id, question_number, correct_option, marks) 
                            VALUES (%s, %s, %s, %s)
                        """, (
                            paper_id, 
                            answer['question_number'], 
                            correct_option,
                            marks
                        ))
            
            conn.commit()
            cursor.close()
            conn.close()
            
            return jsonify({
                "success": True,
                "message": "Answer key saved successfully",
                "paper_id": paper_id,
                "answers_count": len(answers)
            })
            
        except Exception as e:
            if conn:
                conn.rollback()
                conn.close()
            return jsonify({"error": f"Database error: {str(e)}"}), 500
            
    except Exception as e:
        return jsonify({"error": f"Upload failed: {str(e)}"}), 500

@app.route('/api/submit-answers', methods=['POST'])
def student_submit_ocr():
    """Student submit answer sheet with OCR - supports multiple images"""
    try:
        # Check for multiple images (new format) or single image (backward compatibility)
        image_files = request.files.getlist('images')
        if not image_files:
            # Fallback to single image for backward compatibility
            if 'image' in request.files:
                image_files = [request.files['image']]
            else:
                return jsonify({"error": "No image files provided"}), 400
        
        paper_id = request.form.get('paper_id')
        roll_number = request.form.get('roll_number', '').strip()
        student_name = request.form.get('student_name', '').strip()
        
        if not paper_id:
            return jsonify({"error": "Paper ID is required"}), 400
        
        try:
            paper_id = int(paper_id)
        except ValueError:
            return jsonify({"error": "Invalid paper ID"}), 400
        
        # Validate all image files
        for i, image_file in enumerate(image_files):
            if not validate_image_file(image_file):
                return jsonify({"error": f"Invalid image file #{i+1}"}), 400
        
        print(f"Processing {len(image_files)} image(s) for paper {paper_id}")
        
        # Process all images and group by student (roll number)
        students_data = {}  # Dictionary to store data for each student
        
        for i, image_file in enumerate(image_files):
            print(f"Processing image {i+1}/{len(image_files)}: {image_file.filename}")
            
            # Process image with OCR
            image_data = ocr_processor.preprocess_image(image_file)
            if not image_data:
                return jsonify({"error": f"Failed to process image #{i+1}"}), 400
            
            # Extract answers using OCR
            ocr_result = ocr_processor.extract_student_answers(image_data)
            
            # Validate OCR result
            is_valid, error_msg = ocr_processor.validate_student_response(ocr_result)
            if not is_valid:
                return jsonify({"error": f"OCR validation failed for image #{i+1}: {error_msg}"}), 400
            
            # Extract student information
            extracted_roll = ocr_result.get('roll_number', '').strip() if ocr_result.get('roll_number') else ''
            extracted_section = ocr_result.get('section', '').strip() if ocr_result.get('section') else ''
            extracted_paper_info = ocr_result.get('paper_info', {})
            
            if not extracted_roll:
                return jsonify({"error": f"Could not extract roll number from image #{i+1}"}), 400
            
            if not validate_roll_number(extracted_roll):
                return jsonify({"error": f"Invalid roll number format in image #{i+1}: {extracted_roll}"}), 400
            
            # Group by student (roll number)
            if extracted_roll not in students_data:
                students_data[extracted_roll] = {
                    'roll_number': extracted_roll,
                    'section': extracted_section,
                    'paper_info': extracted_paper_info,
                    'answers': {},
                    'image_count': 0
                }
            
            # Update student data
            student_data = students_data[extracted_roll]
            student_data['image_count'] += 1
            
            # Merge paper info (prefer non-null values)
            if extracted_paper_info:
                for key, value in extracted_paper_info.items():
                    if value and not student_data['paper_info'].get(key):
                        student_data['paper_info'][key] = value
            
            # Combine answers from this image for this student
            for answer in ocr_result.get('answers', []):
                question_num = answer.get('question_number')
                if question_num:
                    # If question already exists for this student, prefer the one with more options
                    if question_num not in student_data['answers']:
                        student_data['answers'][question_num] = answer
                    else:
                        # Keep the answer with more selected options (likely more complete)
                        existing_options = student_data['answers'][question_num].get('selected_options', [])
                        new_options = answer.get('selected_options', [])
                        if len(new_options) > len(existing_options):
                            student_data['answers'][question_num] = answer
                            print(f"Updated answer for Q{question_num} for student {extracted_roll} from image {i+1}")
        
        print(f"Found {len(students_data)} unique students in {len(image_files)} images")
        
        # Process each student's data
        all_submissions = []
        for roll_number, student_data in students_data.items():
            print(f"Processing student {roll_number} with data: {student_data.keys()}")
            # Convert answers to list format
            student_answers = list(student_data['answers'].values())
            student_answers.sort(key=lambda x: x.get('question_number', 0))
            
            print(f"Student {roll_number}: {len(student_answers)} questions from {student_data['image_count']} images")
            
            # Convert old format to new format for consistency
            processed_answers = []
            for answer in student_answers:
                if 'selected_options' in answer:
                    # New format - already has selected_options
                    processed_answers.append(answer)
                elif 'selected_option' in answer:
                    # Old format - convert to new format
                    processed_answers.append({
                        'question_number': answer['question_number'],
                        'selected_options': [answer['selected_option']]
                    })
            
            # Create submission data for this student
            submission_data = {
                'roll_number': roll_number,
                'section': student_data['section'],
                'paper_info': student_data['paper_info'],
                'answers': processed_answers,
                'image_count': student_data['image_count']
            }
            all_submissions.append(submission_data)
            print(f"Added submission for student {roll_number}: {len(submission_data['answers'])} answers")
        
        print(f"Total submissions prepared: {len(all_submissions)}")
        for i, sub in enumerate(all_submissions):
            print(f"Submission {i}: Roll {sub['roll_number']}, {len(sub['answers'])} answers")
        
        # Return extracted info for teacher confirmation (multiple students)
        response_data = {
            "success": True,
            "multiple_students": True,
            "students_count": len(students_data),
            "extracted_info": {
                "students": all_submissions,
                "total_images": len(image_files),
                "paper_info": all_submissions[0]['paper_info'] if all_submissions else {}
            },
            "paper_id": paper_id,
            "requires_confirmation": True
        }
        
        print(f"Returning response with {len(response_data['extracted_info']['students'])} students")
        return jsonify(response_data)
        
    except Exception as e:
        return jsonify({"error": f"Submission failed: {str(e)}"}), 500

@app.route('/api/confirm-submission', methods=['POST'])
def confirm_submission():
    """Confirm and save student submission after teacher verification"""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({"error": "No data provided"}), 400
        
        paper_id = data.get('paper_id')
        
        # Check if this is a multiple students submission
        if data.get('multiple_students'):
            students_data = data.get('students', [])
            if not students_data:
                return jsonify({"error": "No student data provided"}), 400
            
            all_results = []
            total_submissions = 0
            
            # Process each student
            for student in students_data:
                roll_number = student.get('roll_number')
                section = student.get('section', '')
                answers = student.get('answers', [])
                images_count = student.get('image_count', 1)
                paper_info = student.get('paper_info', {})
                student_name = student.get('student_name', '')
                
                if not roll_number or not answers:
                    continue  # Skip invalid student data
                
                # Save submission and calculate results for this student
                result = save_student_submission(paper_id, roll_number, section, student_name, answers, images_count, paper_info)
                
                if 'error' not in result:
                    all_results.append({
                        "roll_number": roll_number,
                        "submission_id": result['submission_id'],
                        "answers_count": len(answers),
                        "images_count": images_count
                    })
                    total_submissions += 1
            
            return jsonify({
                "success": True,
                "message": f"Successfully processed {total_submissions} students",
                "multiple_students": True,
                "submissions": all_results,
                "total_students": total_submissions
            })
        
        else:
            # Single student submission (backward compatibility)
            roll_number = data.get('roll_number')
            section = data.get('section')
            student_name = data.get('student_name', '')
            answers = data.get('answers', [])
            images_count = data.get('images_count', 1)
            paper_info = data.get('paper_info', {})
            
            if not paper_id or not roll_number or not answers:
                return jsonify({"error": "Missing required data"}), 400
            
            # Save submission and calculate results
            result = save_student_submission(paper_id, roll_number, section, student_name, answers, images_count, paper_info)
            
            if 'error' in result:
                return jsonify(result), 500
            
            return jsonify({
                "success": True,
                "message": "Answer sheet submitted successfully",
                "submission_id": result['submission_id'],
                "roll_number": roll_number,
                "section": section,
                "extracted_answers": f"{len(answers)} answers processed"
            })
        
    except Exception as e:
        return jsonify({"error": f"Confirmation failed: {str(e)}"}), 500

@app.route('/api/student/submit-manual', methods=['POST'])
def student_submit_manual():
    """Student submit answers manually"""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({"error": "No data provided"}), 400
        
        paper_id = data.get('paper_id')
        roll_number = data.get('roll_number', '').strip()
        student_name = data.get('student_name', '').strip()
        answers = data.get('answers', [])
        
        if not paper_id:
            return jsonify({"error": "Paper ID is required"}), 400
        
        try:
            paper_id = int(paper_id)
        except ValueError:
            return jsonify({"error": "Invalid paper ID"}), 400
        
        if not roll_number:
            return jsonify({"error": "Roll number is required"}), 400
        
        if not validate_roll_number(roll_number):
            return jsonify({"error": "Invalid roll number format"}), 400
        
        if not answers or not isinstance(answers, list):
            return jsonify({"error": "Answers array is required"}), 400
        
        # Validate answers format
        for i, answer in enumerate(answers):
            if not isinstance(answer, dict):
                return jsonify({"error": f"Answer {i+1} must be an object"}), 400
            
            # Check for new format (selected_options) or old format (selected_option)
            if 'selected_options' in answer:
                required_fields = ['question_number', 'selected_options']
            elif 'selected_option' in answer:
                required_fields = ['question_number', 'selected_option']
            else:
                return jsonify({"error": f"Answer {i+1} missing selected options"}), 400
            
            if not all(field in answer for field in required_fields):
                return jsonify({"error": f"Answer {i+1} missing required fields"}), 400
            
            if not isinstance(answer['question_number'], int) or answer['question_number'] <= 0:
                return jsonify({"error": f"Invalid question number in answer {i+1}"}), 400
            
            # Validate selected_options if present
            if 'selected_options' in answer:
                if not isinstance(answer['selected_options'], list) or len(answer['selected_options']) == 0:
                    return jsonify({"error": f"Answer {i+1} selected_options must be non-empty array"}), 400
        
        # Save submission and calculate results
        result = save_student_submission(paper_id, roll_number, student_name, answers)
        
        if 'error' in result:
            return jsonify(result), 500
        
        return jsonify({
            "success": True,
            "message": "Answers submitted successfully",
            "submission_id": result['submission_id'],
            "results": result['results']
        })
        
    except Exception as e:
        return jsonify({"error": f"Submission failed: {str(e)}"}), 500

def save_student_submission(paper_id, roll_number, section, student_name, answers, images_count=1, paper_info=None):
    """Save student submission and calculate results"""
    conn = get_db_connection()
    if not conn:
        return {"error": "Database connection failed"}
    
    try:
        cursor = conn.cursor()
        
        # Check if paper exists
        cursor.execute("SELECT * FROM question_papers WHERE id = %s", (paper_id,))
        paper = cursor.fetchone()
        if not paper:
            cursor.close()
            conn.close()
            return {"error": "Question paper not found"}
        
        # Check if student already submitted - if yes, delete previous submission to allow re-submission
        cursor.execute("""
            SELECT id FROM student_submissions 
            WHERE paper_id = %s AND roll_number = %s
        """, (paper_id, roll_number))
        
        existing = cursor.fetchone()
        if existing:
            # Delete previous submission and related data
            submission_id = existing['id']
            
            # Delete in correct order due to foreign key constraints
            cursor.execute("DELETE FROM results WHERE submission_id = %s", (submission_id,))
            cursor.execute("DELETE FROM student_answers WHERE submission_id = %s", (submission_id,))
            cursor.execute("DELETE FROM student_submissions WHERE id = %s", (submission_id,))
        
        # Use the images_count and paper_info parameters passed to the function
        paper_info = paper_info or {}
        
        try:
            cursor.execute("""
                INSERT INTO student_submissions (
                    paper_id, roll_number, section, student_name, images_count,
                    extracted_paper_name, extracted_subject, extracted_date, 
                    extracted_duration, extracted_total_marks, extracted_class_grade
                ) 
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING id
            """, (
                paper_id, roll_number, section, student_name, images_count,
                paper_info.get('paper_name'), paper_info.get('subject'), paper_info.get('date'),
                paper_info.get('duration'), paper_info.get('total_marks'), paper_info.get('class_grade')
            ))
        except Exception as e:
            # Fallback to old schema if new columns don't exist
            print(f"Using fallback submission insert due to: {e}")
            try:
                cursor.execute("""
                    INSERT INTO student_submissions (paper_id, roll_number, section, student_name, images_count) 
                    VALUES (%s, %s, %s, %s, %s) RETURNING id
                """, (paper_id, roll_number, section, student_name, images_count))
            except Exception as e2:
                # Final fallback to basic schema
                print(f"Using basic fallback submission insert due to: {e2}")
                cursor.execute("""
                    INSERT INTO student_submissions (paper_id, roll_number, section, student_name) 
                    VALUES (%s, %s, %s, %s) RETURNING id
                """, (paper_id, roll_number, section, student_name))
        
        submission_id = cursor.fetchone()['id']
        
        # Insert student answers
        for answer in answers:
            # Handle both new format (selected_options) and old format (selected_option)
            if 'selected_options' in answer:
                selected_options = clean_multiple_options(answer['selected_options'])
                selected_option = selected_options[0] if selected_options else None  # For backward compatibility
            else:
                # Old format
                selected_option = clean_option(answer.get('selected_option', ''))
                selected_options = [selected_option] if selected_option else []
            
            # Try to insert with new schema, fallback to old schema if needed
            try:
                cursor.execute("""
                    INSERT INTO student_answers (submission_id, question_number, selected_option, selected_options) 
                    VALUES (%s, %s, %s, %s)
                """, (
                    submission_id, 
                    answer['question_number'], 
                    selected_option,
                    selected_options
                ))
            except Exception as e:
                # Fallback to old schema if new column doesn't exist
                print(f"Using fallback student answer insert due to: {e}")
                cursor.execute("""
                    INSERT INTO student_answers (submission_id, question_number, selected_option) 
                    VALUES (%s, %s, %s)
                """, (
                    submission_id, 
                    answer['question_number'], 
                    selected_option
                ))
        
        # Calculate results - handle both old and new schema
        try:
            # Try new schema first
            cursor.execute("""
                SELECT ca.question_number, ca.correct_option, ca.correct_options, ca.marks, ca.question_type,
                       sa.selected_option, sa.selected_options
                FROM correct_answers ca
                LEFT JOIN student_answers sa ON ca.question_number = sa.question_number 
                    AND sa.submission_id = %s
                WHERE ca.paper_id = %s
                ORDER BY ca.question_number
            """, (submission_id, paper_id))
        except Exception as e:
            # Fallback to old schema if new columns don't exist
            print(f"Using fallback query due to: {e}")
            cursor.execute("""
                SELECT ca.question_number, ca.correct_option, ca.marks,
                       sa.selected_option,
                       NULL as correct_options, NULL as question_type, NULL as selected_options
                FROM correct_answers ca
                LEFT JOIN student_answers sa ON ca.question_number = sa.question_number 
                    AND sa.submission_id = %s
                WHERE ca.paper_id = %s
                ORDER BY ca.question_number
            """, (submission_id, paper_id))
        
        question_results = cursor.fetchall()
        
        total_questions = len(question_results)
        correct_count = 0
        total_marks = 0
        question_wise_results = []
        
        for result in question_results:
            # Get correct options (handle both new array format and old single format)
            if result['correct_options']:
                # New format - array of correct options
                correct_options = parse_options_string(result['correct_options']) if isinstance(result['correct_options'], str) else result['correct_options']
            else:
                # Fallback to old format
                correct_options = [clean_option(result['correct_option'])] if result['correct_option'] else []
            
            # Get student's selected options (handle both new array format and old single format)
            if result['selected_options']:
                # New format - array of selected options
                selected_options = parse_options_string(result['selected_options']) if isinstance(result['selected_options'], str) else result['selected_options']
            else:
                # Fallback to old format
                selected_options = [clean_option(result['selected_option'])] if result['selected_option'] else []
            
            # Clean the options
            correct_options = clean_multiple_options(correct_options)
            selected_options = clean_multiple_options(selected_options)
            
            # Use new partial marking system
            marking_result = calculate_partial_marks(selected_options, correct_options, result['marks'])
            
            # Update counters based on new marking system
            if marking_result['is_fully_correct']:
                correct_count += 1
            
            total_marks += marking_result['marks_awarded']
            
            # Determine status based on new marking system
            if not selected_options:
                status_value = "not_answered"
            elif marking_result['is_fully_correct']:
                status_value = "correct"
            elif marking_result['is_partially_correct']:
                status_value = "partial"
            else:
                status_value = "incorrect"
            
            # Format options for display
            correct_display = ', '.join(correct_options) if correct_options else 'N/A'
            selected_display = ', '.join(selected_options) if selected_options else 'N/A'
            
            question_wise_results.append({
                "question_number": result['question_number'],
                "correct_option": correct_display,  # For backward compatibility
                "correct_options": correct_options,  # New field for multiple options
                "selected_option": selected_display,  # For backward compatibility
                "selected_options": selected_options,  # New field for multiple options
                "is_correct": marking_result['is_fully_correct'],
                "is_partially_correct": marking_result['is_partially_correct'],
                "has_wrong_options": marking_result['has_wrong_options'],
                "marks": marking_result['marks_awarded'],
                "total_marks": result['marks'],  # Total marks available for this question
                "status": status_value,
                "question_type": result.get('question_type', 'single'),
                "marking_explanation": marking_result['explanation']
            })
        
        # Calculate percentage
        max_marks = sum(r['marks'] for r in question_results)
        percentage = calculate_percentage(total_marks, max_marks)
        
        # Insert results
        cursor.execute("""
            INSERT INTO results (submission_id, total_questions, correct_answers, total_marks, percentage) 
            VALUES (%s, %s, %s, %s, %s)
        """, (submission_id, total_questions, correct_count, total_marks, percentage))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return {
            "submission_id": submission_id,
            "results": {
                "total_questions": total_questions,
                "correct_answers": correct_count,
                "total_marks": total_marks,
                "max_marks": max_marks,
                "percentage": float(percentage),
                "question_wise": question_wise_results
            }
        }
        
    except Exception as e:
        if conn:
            conn.rollback()
            conn.close()
        return {"error": f"Database error: {str(e)}"}

@app.route('/api/papers/<int:paper_id>/answers', methods=['GET'])
def get_paper_answers(paper_id):
    """Get answers for a specific paper"""
    conn = get_db_connection()
    if not conn:
        return jsonify({"error": "Database connection failed"}), 500
    
    try:
        cursor = conn.cursor()
        
        # Get paper info
        cursor.execute("SELECT * FROM question_papers WHERE id = %s", (paper_id,))
        paper = cursor.fetchone()
        if not paper:
            cursor.close()
            conn.close()
            return jsonify({"error": "Paper not found"}), 404
        
        # Get answers - handle both old and new schema
        try:
            # Try new schema first
            cursor.execute("""
                SELECT question_number, correct_option, correct_options, marks, question_type 
                FROM correct_answers 
                WHERE paper_id = %s 
                ORDER BY question_number
            """, (paper_id,))
        except Exception as e:
            # Fallback to old schema if new columns don't exist
            print(f"Using fallback paper answers query due to: {e}")
            cursor.execute("""
                SELECT question_number, correct_option, marks,
                       NULL as correct_options, NULL as question_type
                FROM correct_answers 
                WHERE paper_id = %s 
                ORDER BY question_number
            """, (paper_id,))
        answers = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            "paper": dict(paper),
            "answers": [dict(answer) for answer in answers]
        })
        
    except Exception as e:
        if conn:
            conn.close()
        return jsonify({"error": f"Failed to get answers: {str(e)}"}), 500

@app.route('/api/papers/<int:paper_id>', methods=['DELETE'])
def delete_paper(paper_id):
    """Delete a paper and all related data"""
    conn = get_db_connection()
    if not conn:
        return jsonify({"error": "Database connection failed"}), 500
    
    try:
        cursor = conn.cursor()
        
        # Check if paper exists
        cursor.execute("SELECT paper_name FROM question_papers WHERE id = %s", (paper_id,))
        paper = cursor.fetchone()
        if not paper:
            cursor.close()
            conn.close()
            return jsonify({"error": "Paper not found"}), 404
        
        # Defensive cleanup in case ON DELETE CASCADE wasn't applied at creation time
        # Delete dependent rows explicitly to avoid FK violations in existing databases
        cursor.execute("DELETE FROM correct_answers WHERE paper_id = %s", (paper_id,))
        cursor.execute("DELETE FROM student_answers WHERE submission_id IN (SELECT id FROM student_submissions WHERE paper_id = %s)", (paper_id,))
        cursor.execute("DELETE FROM results WHERE submission_id IN (SELECT id FROM student_submissions WHERE paper_id = %s)", (paper_id,))
        cursor.execute("DELETE FROM student_submissions WHERE paper_id = %s", (paper_id,))
        # Finally delete the paper
        cursor.execute("DELETE FROM question_papers WHERE id = %s", (paper_id,))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({"message": f"Paper '{paper['paper_name']}' deleted successfully"})
        
    except Exception as e:
        if conn:
            conn.rollback()
            conn.close()
        return jsonify({"error": f"Failed to delete paper: {str(e)}"}), 500

@app.route('/api/results/<roll_number>', methods=['GET'])
def search_results(roll_number):
    """Search results by roll number"""
    if not roll_number.strip():
        return jsonify({"error": "Roll number is required"}), 400
    
    conn = get_db_connection()
    if not conn:
        return jsonify({"error": "Database connection failed"}), 500
    
    try:
        cursor = conn.cursor()
        
        # Try to get results with images_count, fallback if column doesn't exist
        try:
            cursor.execute("""
                SELECT ss.id as submission_id, ss.roll_number, ss.student_name, 
                       qp.paper_name, r.total_questions, r.correct_answers, 
                       r.total_marks, r.percentage, ss.submitted_at, ss.images_count
                FROM student_submissions ss
                JOIN question_papers qp ON ss.paper_id = qp.id
                JOIN results r ON ss.id = r.submission_id
                WHERE ss.roll_number = %s
                ORDER BY ss.submitted_at DESC
            """, (roll_number,))
        except Exception as e:
            # Fallback without images_count
            print(f"Using fallback results query due to: {e}")
            cursor.execute("""
                SELECT ss.id as submission_id, ss.roll_number, ss.student_name, 
                       qp.paper_name, r.total_questions, r.correct_answers, 
                       r.total_marks, r.percentage, ss.submitted_at, 1 as images_count
                FROM student_submissions ss
                JOIN question_papers qp ON ss.paper_id = qp.id
                JOIN results r ON ss.id = r.submission_id
                WHERE ss.roll_number = %s
                ORDER BY ss.submitted_at DESC
            """, (roll_number,))
        
        results = cursor.fetchall()
        cursor.close()
        conn.close()
        
        return jsonify([dict(result) for result in results])
        
    except Exception as e:
        if conn:
            conn.close()
        return jsonify({"error": f"Search failed: {str(e)}"}), 500

@app.route('/api/results/analytics', methods=['GET'])
def get_analytics():
    """Get system analytics"""
    conn = get_db_connection()
    if not conn:
        return jsonify({"error": "Database connection failed"}), 500
    
    try:
        cursor = conn.cursor()
        
        # Overview stats
        cursor.execute("SELECT COUNT(*) as count FROM question_papers")
        total_papers = cursor.fetchone()['count']
        
        cursor.execute("SELECT COUNT(*) as count FROM student_submissions")
        total_submissions = cursor.fetchone()['count']
        
        cursor.execute("SELECT COUNT(DISTINCT roll_number) as count FROM student_submissions")
        unique_students = cursor.fetchone()['count']
        
        cursor.execute("SELECT AVG(percentage) as avg_percentage FROM results")
        avg_result = cursor.fetchone()
        avg_percentage = round(float(avg_result['avg_percentage']) if avg_result['avg_percentage'] else 0, 1)
        
        # Top performers
        cursor.execute("""
            SELECT ss.roll_number, ss.student_name, qp.paper_name, r.percentage
            FROM results r
            JOIN student_submissions ss ON r.submission_id = ss.id
            JOIN question_papers qp ON ss.paper_id = qp.id
            ORDER BY r.percentage DESC
            LIMIT 10
        """)
        top_performers = cursor.fetchall()
        
        # Grade distribution (use subquery to avoid GROUP BY on percentage)
        cursor.execute("""
            SELECT grade, COUNT(*) AS count
            FROM (
                SELECT CASE 
                    WHEN percentage >= 90 THEN 'A+'
                    WHEN percentage >= 80 THEN 'A'
                    WHEN percentage >= 70 THEN 'B+'
                    WHEN percentage >= 60 THEN 'B'
                    WHEN percentage >= 50 THEN 'C'
                    ELSE 'F'
                END AS grade
                FROM results
            ) AS graded
            GROUP BY grade
            ORDER BY CASE grade
                WHEN 'A+' THEN 1
                WHEN 'A' THEN 2
                WHEN 'B+' THEN 3
                WHEN 'B' THEN 4
                WHEN 'C' THEN 5
                ELSE 6
            END
        """)
        grade_distribution = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            "overview": {
                "total_papers": total_papers,
                "total_submissions": total_submissions,
                "unique_students": unique_students,
                "average_performance": avg_percentage
            },
            "top_performers": [dict(performer) for performer in top_performers],
            "grade_distribution": [dict(grade) for grade in grade_distribution]
        })
        
    except Exception as e:
        if conn:
            conn.close()
        return jsonify({"error": f"Failed to get analytics: {str(e)}"}), 500

@app.route('/api/submissions', methods=['GET'])
def get_all_submissions():
    """Get all submissions"""
    conn = get_db_connection()
    if not conn:
        return jsonify({"error": "Database connection failed"}), 500
    
    try:
        cursor = conn.cursor()
        
        # Get pagination parameters
        limit = int(request.args.get('limit', 50))
        offset = int(request.args.get('offset', 0))
        
        # Try to get submissions with images_count, fallback if column doesn't exist
        try:
            cursor.execute("""
                SELECT ss.id, ss.roll_number, ss.student_name, qp.paper_name, 
                       r.total_questions, r.correct_answers, r.percentage, ss.submitted_at, ss.images_count
                FROM student_submissions ss
                JOIN question_papers qp ON ss.paper_id = qp.id
                JOIN results r ON ss.id = r.submission_id
                ORDER BY ss.submitted_at DESC
                LIMIT %s OFFSET %s
            """, (limit, offset))
        except Exception as e:
            # Fallback without images_count
            print(f"Using fallback all submissions query due to: {e}")
            cursor.execute("""
                SELECT ss.id, ss.roll_number, ss.student_name, qp.paper_name, 
                       r.total_questions, r.correct_answers, r.percentage, ss.submitted_at, 1 as images_count
                FROM student_submissions ss
                JOIN question_papers qp ON ss.paper_id = qp.id
                JOIN results r ON ss.id = r.submission_id
                ORDER BY ss.submitted_at DESC
                LIMIT %s OFFSET %s
            """, (limit, offset))
        
        submissions = cursor.fetchall()
        cursor.close()
        conn.close()
        
        return jsonify([dict(submission) for submission in submissions])
        
    except Exception as e:
        if conn:
            conn.close()
        return jsonify({"error": f"Failed to get submissions: {str(e)}"}), 500

@app.route('/api/results/detailed/<int:submission_id>', methods=['GET'])
def get_detailed_results(submission_id):
    """Get detailed results for a specific submission including question-wise breakdown"""
    conn = get_db_connection()
    if not conn:
        return jsonify({"error": "Database connection failed"}), 500
    try:
        cursor = conn.cursor()
        # Fetch submission basic data
        cursor.execute(
            """
            SELECT ss.id as submission_id, ss.roll_number, ss.student_name, ss.submitted_at,
                   qp.paper_name,
                   r.total_questions, r.correct_answers, r.total_marks, r.percentage
            FROM student_submissions ss
            JOIN question_papers qp ON ss.paper_id = qp.id
            JOIN results r ON ss.id = r.submission_id
            WHERE ss.id = %s
            """,
            (submission_id,)
        )
        submission = cursor.fetchone()
        if not submission:
            cursor.close()
            conn.close()
            return jsonify({"error": "Submission not found"}), 404

        # Fetch question-wise mapping - handle both old and new schema
        try:
            # Try new schema first
            cursor.execute(
                """
                SELECT ca.question_number, ca.correct_option, ca.correct_options, ca.marks, ca.question_type,
                       sa.selected_option, sa.selected_options
                FROM correct_answers ca
                LEFT JOIN student_answers sa
                  ON sa.submission_id = %s AND sa.question_number = ca.question_number
                WHERE ca.paper_id = (
                    SELECT paper_id FROM student_submissions WHERE id = %s
                )
                ORDER BY ca.question_number
                """,
                (submission_id, submission_id)
            )
        except Exception as e:
            # Fallback to old schema if new columns don't exist
            print(f"Using fallback detailed query due to: {e}")
            cursor.execute(
                """
                SELECT ca.question_number, ca.correct_option, ca.marks,
                       sa.selected_option,
                       NULL as correct_options, NULL as question_type, NULL as selected_options
                FROM correct_answers ca
                LEFT JOIN student_answers sa
                  ON sa.submission_id = %s AND sa.question_number = ca.question_number
                WHERE ca.paper_id = (
                    SELECT paper_id FROM student_submissions WHERE id = %s
                )
                ORDER BY ca.question_number
                """,
                (submission_id, submission_id)
            )
        rows = cursor.fetchall()

        questions = []
        correct = 0
        incorrect = 0
        not_answered = 0
        for row in rows:
            # Get correct options (handle both new array format and old single format)
            if row['correct_options']:
                # New format - array of correct options
                correct_options = parse_options_string(row['correct_options']) if isinstance(row['correct_options'], str) else row['correct_options']
            else:
                # Fallback to old format
                correct_options = [clean_option(row['correct_option'])] if row['correct_option'] else []
            
            # Get student's selected options (handle both new array format and old single format)
            if row['selected_options']:
                # New format - array of selected options
                selected_options = parse_options_string(row['selected_options']) if isinstance(row['selected_options'], str) else row['selected_options']
            else:
                # Fallback to old format
                selected_options = [clean_option(row['selected_option'])] if row['selected_option'] else []
            
            # Clean the options
            correct_options = clean_multiple_options(correct_options)
            selected_options = clean_multiple_options(selected_options)
            
            # Use new partial marking system
            marking_result = calculate_partial_marks(selected_options, correct_options, row['marks'])
            
            # Update counters based on new marking system
            if not selected_options:
                not_answered += 1
            elif marking_result['is_fully_correct']:
                correct += 1
            elif marking_result['is_partially_correct']:
                # Count partial as separate category, but for now add to correct for percentage calculation
                correct += 1  # This might need adjustment based on your requirements
            else:
                incorrect += 1
            
            # Determine status based on new marking system
            if not selected_options:
                status_value = "not_answered"
            elif marking_result['is_fully_correct']:
                status_value = "correct"
            elif marking_result['is_partially_correct']:
                status_value = "partial"
            else:
                status_value = "incorrect"
            
            # Format options for display
            correct_display = ', '.join(correct_options) if correct_options else 'N/A'
            selected_display = ', '.join(selected_options) if selected_options else 'N/A'
            
            questions.append({
                'question_number': row['question_number'],
                'correct_option': correct_display,  # For backward compatibility
                'correct_options': correct_options,  # New field for multiple options
                'selected_option': selected_display,  # For backward compatibility
                'selected_options': selected_options,  # New field for multiple options
                'status': status_value,
                'is_partially_correct': marking_result['is_partially_correct'],
                'has_wrong_options': marking_result['has_wrong_options'],
                'marks': marking_result['marks_awarded'],
                'total_marks': row['marks'],
                'question_type': row.get('question_type', 'single'),
                'marking_explanation': marking_result['explanation']
            })

        statistics = {
            'correct': correct,
            'incorrect': incorrect,
            'not_answered': not_answered,
            'total_questions': submission['total_questions'],
            'percentage': float(submission['percentage']) if submission['percentage'] is not None else 0.0
        }

        response = {
            'submission': dict(submission),
            'questions': questions,
            'statistics': statistics
        }

        cursor.close()
        conn.close()
        return jsonify(response)
    except Exception as e:
        if conn:
            conn.close()
        return jsonify({"error": f"Failed to get detailed results: {str(e)}"}), 500

@app.route('/api/export/<int:paper_id>', methods=['GET'])
def export_paper_results(paper_id):
    """Export paper results to Excel"""
    conn = get_db_connection()
    if not conn:
        return jsonify({"error": "Database connection failed"}), 500
    
    try:
        cursor = conn.cursor()
        
        # Get paper information
        cursor.execute("SELECT paper_name, total_questions FROM question_papers WHERE id = %s", (paper_id,))
        paper = cursor.fetchone()
        if not paper:
            cursor.close()
            conn.close()
            return jsonify({"error": "Paper not found"}), 404
        
        # Get all submissions for this paper with results (handling duplicates by taking latest)
        cursor.execute("""
            SELECT DISTINCT ON (ss.roll_number) 
                   ss.roll_number, 
                   ss.student_name, 
                   r.total_marks,
                   r.percentage,
                   r.correct_answers,
                   r.total_questions,
                   ss.submitted_at,
                   CASE 
                       WHEN r.percentage >= 90 THEN 'A+'
                       WHEN r.percentage >= 80 THEN 'A'
                       WHEN r.percentage >= 70 THEN 'B+'
                       WHEN r.percentage >= 60 THEN 'B'
                       WHEN r.percentage >= 50 THEN 'C'
                       ELSE 'F'
                   END as grade
            FROM student_submissions ss
            JOIN results r ON ss.id = r.submission_id
            WHERE ss.paper_id = %s
            ORDER BY ss.roll_number, ss.submitted_at DESC
        """, (paper_id,))
        
        submissions = cursor.fetchall()
        
        # Get max possible marks for the paper
        cursor.execute("SELECT SUM(marks) as max_marks FROM correct_answers WHERE paper_id = %s", (paper_id,))
        max_marks_result = cursor.fetchone()
        max_marks = max_marks_result['max_marks'] if max_marks_result and max_marks_result['max_marks'] else 0
        
        cursor.close()
        conn.close()
        
        if not submissions:
            return jsonify({"error": "No submissions found for this paper"}), 404
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Add title and paper info
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = f"Results for {paper['paper_name']}"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = center_alignment
        
        ws.merge_cells('A2:H2')
        info_cell = ws['A2']
        info_cell.value = f"Total Questions: {paper['total_questions']} | Max Marks: {max_marks} | Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        info_cell.alignment = center_alignment
        
        # Add headers
        headers = ['S.No.', 'Roll Number', 'Student Name', 'Marks Obtained', 'Max Marks', 'Percentage (%)', 'Grade', 'Submitted At']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
        
        # Add data rows
        for idx, submission in enumerate(submissions, 1):
            row = idx + 4
            
            # Serial number
            ws.cell(row=row, column=1, value=idx).border = border
            ws.cell(row=row, column=1).alignment = center_alignment
            
            # Roll number
            ws.cell(row=row, column=2, value=submission['roll_number']).border = border
            ws.cell(row=row, column=2).alignment = center_alignment
            
            # Student name
            student_name = submission['student_name'] if submission['student_name'] else 'Not provided'
            ws.cell(row=row, column=3, value=student_name).border = border
            
            # Marks obtained
            ws.cell(row=row, column=4, value=submission['total_marks']).border = border
            ws.cell(row=row, column=4).alignment = center_alignment
            
            # Max marks
            ws.cell(row=row, column=5, value=max_marks).border = border
            ws.cell(row=row, column=5).alignment = center_alignment
            
            # Percentage
            ws.cell(row=row, column=6, value=f"{submission['percentage']:.1f}").border = border
            ws.cell(row=row, column=6).alignment = center_alignment
            
            # Grade
            grade_cell = ws.cell(row=row, column=7, value=submission['grade'])
            grade_cell.border = border
            grade_cell.alignment = center_alignment
            
            # Color code grades
            if submission['grade'] in ['A+', 'A']:
                grade_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            elif submission['grade'] in ['B+', 'B']:
                grade_cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
            elif submission['grade'] == 'C':
                grade_cell.fill = PatternFill(start_color="FFB347", end_color="FFB347", fill_type="solid")
            else:  # F
                grade_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            
            # Submitted at
            submitted_at = submission['submitted_at'].strftime('%Y-%m-%d %H:%M:%S') if submission['submitted_at'] else 'N/A'
            ws.cell(row=row, column=8, value=submitted_at).border = border
            ws.cell(row=row, column=8).alignment = center_alignment
        
        # Add summary statistics
        summary_row = len(submissions) + 6
        ws.merge_cells(f'A{summary_row}:H{summary_row}')
        summary_cell = ws[f'A{summary_row}']
        summary_cell.value = "Summary Statistics"
        summary_cell.font = Font(bold=True, size=14)
        summary_cell.alignment = center_alignment
        
        # Calculate statistics
        total_students = len(submissions)
        avg_marks = sum(s['total_marks'] for s in submissions) / total_students if total_students > 0 else 0
        avg_percentage = sum(float(s['percentage']) for s in submissions) / total_students if total_students > 0 else 0
        
        grade_counts = {}
        for submission in submissions:
            grade = submission['grade']
            grade_counts[grade] = grade_counts.get(grade, 0) + 1
        
        # Add statistics
        stats_start_row = summary_row + 2
        stats = [
            ['Total Students:', total_students],
            ['Average Marks:', f"{avg_marks:.1f}/{max_marks}"],
            ['Average Percentage:', f"{avg_percentage:.1f}%"],
            ['Grade Distribution:', '']
        ]
        
        for grade in ['A+', 'A', 'B+', 'B', 'C', 'F']:
            if grade in grade_counts:
                stats.append([f'  {grade}:', f"{grade_counts[grade]} ({grade_counts[grade]/total_students*100:.1f}%)"])
        
        for i, (label, value) in enumerate(stats):
            ws.cell(row=stats_start_row + i, column=1, value=label).font = Font(bold=True)
            ws.cell(row=stats_start_row + i, column=2, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Create filename
        safe_paper_name = sanitize_filename(paper['paper_name'])
        filename = f"{safe_paper_name}_Results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        response = send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return response
        
    except Exception as e:
        if conn:
            conn.close()
        return jsonify({"error": f"Export failed: {str(e)}"}), 500

@app.route('/api/papers/<int:paper_id>/submissions-preview', methods=['GET'])
def get_submissions_preview(paper_id):
    """Get preview of submissions for a paper before export"""
    conn = get_db_connection()
    if not conn:
        return jsonify({"error": "Database connection failed"}), 500
    
    try:
        cursor = conn.cursor()
        
        # Get paper information
        cursor.execute("SELECT paper_name, total_questions FROM question_papers WHERE id = %s", (paper_id,))
        paper = cursor.fetchone()
        if not paper:
            cursor.close()
            conn.close()
            return jsonify({"error": "Paper not found"}), 404
        
        # Get submissions preview (handling duplicates by taking latest)
        cursor.execute("""
            SELECT DISTINCT ON (ss.roll_number) 
                   ss.roll_number, 
                   ss.student_name, 
                   r.total_marks,
                   r.percentage
            FROM student_submissions ss
            JOIN results r ON ss.id = r.submission_id
            WHERE ss.paper_id = %s
            ORDER BY ss.roll_number, ss.submitted_at DESC
            LIMIT 50
        """, (paper_id,))
        
        submissions = cursor.fetchall()
        
        # Get max possible marks for the paper
        cursor.execute("SELECT SUM(marks) as max_marks FROM correct_answers WHERE paper_id = %s", (paper_id,))
        max_marks_result = cursor.fetchone()
        max_marks = max_marks_result['max_marks'] if max_marks_result and max_marks_result['max_marks'] else 0
        
        cursor.close()
        conn.close()
        
        return jsonify({
            "paper_name": paper['paper_name'],
            "total_questions": paper['total_questions'],
            "max_marks": max_marks,
            "submissions": [dict(submission) for submission in submissions],
            "total_submissions": len(submissions)
        })
        
    except Exception as e:
        if conn:
            conn.close()
        return jsonify({"error": f"Preview failed: {str(e)}"}), 500

@app.route('/health')
def health_check():
    """Health check endpoint"""
    return jsonify({"status": "healthy", "service": "flask-ocr-grading"})

@app.route('/migrate-db')
def migrate_database():
    """Add section column to student_submissions table"""
    conn = get_db_connection()
    if not conn:
        return "Database connection failed", 500
    
    try:
        cursor = conn.cursor()
        
        # Check if column exists
        cursor.execute("""
            SELECT column_name 
            FROM information_schema.columns 
            WHERE table_name='student_submissions' AND column_name='section';
        """)
        
        if cursor.fetchone():
            return "Section column already exists"
        
        # Add section column
        cursor.execute("ALTER TABLE student_submissions ADD COLUMN section VARCHAR(10);")
        
        # Create index
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_student_submissions_section ON student_submissions(section);")
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return "✅ Migration completed successfully! Section column added."
        
    except Exception as e:
        if conn:
            conn.rollback()
            conn.close()
        return f"Migration failed: {str(e)}", 500

@app.route('/api/debug/test-multi-student', methods=['GET'])
def debug_multi_student():
    """Debug endpoint to test multi-student functionality"""
    return jsonify({
        "message": "Multi-student debug endpoint",
        "features": {
            "multi_image_upload": True,
            "multi_student_detection": True,
            "paper_info_extraction": True,
            "separate_submissions": True
        },
        "expected_flow": [
            "1. Upload multiple images",
            "2. OCR extracts roll numbers",
            "3. Group by unique roll number",
            "4. Show confirmation for all students",
            "5. Create separate submissions"
        ]
    })

@app.route('/api/debug/ocr-marks', methods=['POST'])
def debug_ocr_marks():
    """Debug endpoint to test OCR mark detection"""
    try:
        if 'image' not in request.files:
            return jsonify({"error": "No image file provided"}), 400
        
        image_file = request.files['image']
        
        if not validate_image_file(image_file):
            return jsonify({"error": "Invalid image file"}), 400
        
        # Process image with OCR
        image_data = ocr_processor.preprocess_image(image_file)
        if not image_data:
            return jsonify({"error": "Failed to process image"}), 400
        
        # Run debug mark detection
        debug_info = ocr_processor.debug_mark_detection(image_data)
        
        # Also run normal extraction for comparison
        ocr_result = ocr_processor.extract_student_answers(image_data)
        
        return jsonify({
            "debug_info": debug_info,
            "ocr_result": ocr_result,
            "message": "Check server logs for detailed debug output"
        })
        
    except Exception as e:
        return jsonify({"error": f"Debug failed: {str(e)}"}), 500

try:
    init_database()
except Exception as e:
    print(f"Failed to initialize database on startup: {e}")

# For Vercel deployment
app.wsgi_app = app.wsgi_app

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug_mode = os.environ.get('VERCEL_ENV') != 'production'
    app.run(debug=debug_mode, host='0.0.0.0', port=port)
